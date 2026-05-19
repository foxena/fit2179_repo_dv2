# COROS FIT parser fix: data messages must skip developer-field bytes, otherwise many Running sessions are missed.
import os, sys, json, zipfile, shutil, math, statistics
from pathlib import Path
from datetime import datetime, timedelta, timezone, date
from zoneinfo import ZoneInfo
from collections import defaultdict, Counter

import pandas as pd
from openpyxl import load_workbook

sys.path.append('/mnt/data')
from fit_minimal import parse_fit, SPORTS

BASE = Path('/mnt/data')
PROJECT = BASE / 'fit2179_running_dv2'
DATA_DIR = PROJECT / 'data'
JS_DIR = PROJECT / 'js'
CSS_DIR = PROJECT / 'css'
SCRIPTS_DIR = PROJECT / 'scripts'

for d in [PROJECT, DATA_DIR, JS_DIR, CSS_DIR, SCRIPTS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

MELB = ZoneInfo('Australia/Melbourne')

# -------------------------
# FIT / COROS processing
# -------------------------
FIT_ZIP = BASE / 'exportSportData_466571768635605001_20260518.zip'
FIT_DIR = BASE / 'coros_fit'
if not FIT_DIR.exists():
    FIT_DIR.mkdir()
    with zipfile.ZipFile(FIT_ZIP) as z:
        z.extractall(FIT_DIR)

def local_iso(dt):
    if dt is None: return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(MELB).isoformat(timespec='seconds')

def local_date_str(dt):
    if dt is None: return None
    if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(MELB).date().isoformat()

def monday_of(d):
    if isinstance(d, str): d = date.fromisoformat(d)
    return (d - timedelta(days=d.weekday())).isoformat()

activities=[]
parse_errors=[]
for p in sorted(FIT_DIR.glob('*.fit')):
    try:
        msgs = parse_fit(str(p))
    except Exception as e:
        parse_errors.append({'file': p.name, 'error': str(e)})
        continue
    sessions=[m for m in msgs if m.get('global')==18]
    records=[m for m in msgs if m.get('global')==20]
    if not sessions:
        continue
    # Most files have one session; if multiple, use each.
    for sidx, s in enumerate(sessions):
        distance_m = s.get('total_distance')
        timer_s = s.get('total_timer_time') or s.get('total_elapsed_time')
        if distance_m is None or timer_s is None or timer_s <= 0 or distance_m <= 0:
            continue
        start_dt = s.get('start_time') or s.get('timestamp')
        if start_dt is None and records:
            start_dt = records[0].get('timestamp')
        dstr = local_date_str(start_dt)
        if not dstr:
            continue
        sport_code = s.get('sport')
        sport_name = SPORTS.get(sport_code, f'Sport {sport_code}')
        # Keep the project aligned with running/foot-based endurance, but keep other activities in data.
        if sport_name in ['Running','Generic Running']:
            activity_group = 'Running'
        elif sport_name in ['Hiking','Walking']:
            activity_group = 'Foot-based'
        else:
            activity_group = 'Other training'
        distance_km = distance_m / 1000
        duration_min = timer_s / 60
        pace = duration_min / distance_km if distance_km else None
        avg_speed_kmh = (distance_km / (duration_min/60)) if duration_min else None
        total_ascent = s.get('total_ascent') or 0
        ascent_per_km = total_ascent / distance_km if distance_km else None
        activities.append({
            'activity_id': p.stem + (f'_{sidx}' if len(sessions)>1 else ''),
            'source_file': p.name,
            'date': dstr,
            'start_time_local': local_iso(start_dt),
            'year': int(dstr[:4]),
            'month': dstr[:7],
            'week_start': monday_of(dstr),
            'weekday': ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][date.fromisoformat(dstr).weekday()],
            'weekday_num': date.fromisoformat(dstr).weekday(),
            'sport': sport_name,
            'activity_group': activity_group,
            'distance_km': round(distance_km, 2),
            'duration_min': round(duration_min, 1),
            'elapsed_min': round((s.get('total_elapsed_time') or timer_s)/60, 1),
            'pace_min_km': round(pace, 2) if pace else None,
            'avg_speed_kmh': round(avg_speed_kmh, 2) if avg_speed_kmh else None,
            'avg_hr': s.get('avg_heart_rate'),
            'max_hr': s.get('max_heart_rate'),
            'elevation_gain_m': total_ascent,
            'elevation_loss_m': s.get('total_descent') or 0,
            'ascent_per_km': round(ascent_per_km, 1) if ascent_per_km is not None else None,
            'avg_cadence': s.get('avg_cadence'),
            'max_cadence': s.get('max_cadence'),
            'calories': s.get('total_calories'),
            'distance_band': '<2 km' if distance_km < 2 else '2–5 km' if distance_km < 5 else '5–10 km' if distance_km < 10 else '10+ km',
            'duration_band': '<30 min' if duration_min < 30 else '30–60 min' if duration_min < 60 else '1–2 hr' if duration_min < 120 else '2+ hr'
        })

# Add effort score: derived and relative, from HR, elevation density and distance.
def zscores(vals):
    nums=[v for v in vals if isinstance(v,(int,float)) and not math.isnan(v)]
    mean=statistics.mean(nums) if nums else 0
    st=statistics.pstdev(nums) if len(nums)>1 else 1
    if st == 0: st=1
    return [(None if v is None else (v-mean)/st) for v in vals]

for metric in ['avg_hr','ascent_per_km','distance_km']:
    zs=zscores([a.get(metric) for a in activities])
    for a,z in zip(activities,zs):
        a[f'z_{metric}']=z
for a in activities:
    score = 50
    if a.get('z_avg_hr') is not None: score += 18*a['z_avg_hr']
    if a.get('z_ascent_per_km') is not None: score += 12*a['z_ascent_per_km']
    if a.get('z_distance_km') is not None: score += 10*a['z_distance_km']
    a['effort_score'] = round(max(5, min(100, score)), 1)
    a.pop('z_avg_hr', None); a.pop('z_ascent_per_km', None); a.pop('z_distance_km', None)

# Keep all summary rows, but flag foot-based ones for main text.
with open(DATA_DIR/'coros_activities.json','w') as f:
    json.dump(activities, f, indent=2)

# Summary tables
adf=pd.DataFrame(activities)
if adf.empty:
    raise RuntimeError('No usable COROS activities found')

# Weekly summary
weekly=(adf.groupby('week_start', as_index=False)
        .agg(total_distance_km=('distance_km','sum'), activities=('activity_id','count'), total_duration_min=('duration_min','sum'), avg_effort=('effort_score','mean')))
weekly['total_distance_km']=weekly['total_distance_km'].round(1)
weekly['total_duration_hr']=(weekly['total_duration_min']/60).round(1)
weekly['avg_effort']=weekly['avg_effort'].round(1)
weekly.to_json(DATA_DIR/'coros_weekly.json', orient='records', indent=2)

monthly=(adf.groupby('month', as_index=False)
         .agg(total_distance_km=('distance_km','sum'), activities=('activity_id','count'), total_duration_min=('duration_min','sum'), avg_pace=('pace_min_km','mean'), avg_hr=('avg_hr','mean'), elevation_gain_m=('elevation_gain_m','sum')))
for col in ['total_distance_km','total_duration_min','avg_pace','avg_hr','elevation_gain_m']:
    monthly[col]=monthly[col].round(1)
monthly['total_duration_hr']=(monthly['total_duration_min']/60).round(1)
monthly.to_json(DATA_DIR/'coros_monthly.json', orient='records', indent=2)

dow=(adf.groupby(['weekday','weekday_num'], as_index=False)
     .agg(activities=('activity_id','count'), total_distance_km=('distance_km','sum'), avg_distance_km=('distance_km','mean')))
for col in ['total_distance_km','avg_distance_km']:
    dow[col]=dow[col].round(1)
dow.sort_values('weekday_num').to_json(DATA_DIR/'coros_dayofweek.json', orient='records', indent=2)

cal=(adf.groupby('date', as_index=False)
     .agg(distance_km=('distance_km','sum'), activities=('activity_id','count'), avg_effort=('effort_score','mean')))
cal['distance_km']=cal['distance_km'].round(1); cal['avg_effort']=cal['avg_effort'].round(1)
cal['week_start']=cal['date'].map(monday_of)
cal['weekday_num']=[date.fromisoformat(d).weekday() for d in cal['date']]
cal['weekday']=[['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][n] for n in cal['weekday_num']]
cal['month']=cal['date'].str.slice(0,7)
cal.to_json(DATA_DIR/'coros_calendar.json', orient='records', indent=2)

sport_summary=(adf.groupby(['sport','activity_group'], as_index=False)
               .agg(activities=('activity_id','count'), distance_km=('distance_km','sum'), duration_min=('duration_min','sum')))
sport_summary['distance_km']=sport_summary['distance_km'].round(1)
sport_summary['duration_hr']=(sport_summary['duration_min']/60).round(1)
sport_summary.to_json(DATA_DIR/'coros_sport_summary.json', orient='records', indent=2)

# Custom matrix: effort level vs distance band
bins=[]
for _,r in adf.iterrows():
    score=r['effort_score']
    level='Low' if score < 40 else 'Moderate' if score < 60 else 'Hard' if score < 80 else 'Very hard'
    bins.append({'distance_band':r['distance_band'], 'effort_level':level, 'activity_id':r['activity_id'], 'distance_km':r['distance_km']})
bdf=pd.DataFrame(bins)
effort_matrix=(bdf.groupby(['distance_band','effort_level'], as_index=False).agg(activities=('activity_id','count'), distance_km=('distance_km','sum')))
effort_matrix['distance_km']=effort_matrix['distance_km'].round(1)
effort_matrix.to_json(DATA_DIR/'coros_effort_matrix.json', orient='records', indent=2)

# KPIs
foot=adf[adf['activity_group'].isin(['Running','Foot-based'])]
kpi_source=foot if not foot.empty else adf
kpis={
    'activity_count': int(len(kpi_source)),
    'total_distance_km': round(float(kpi_source['distance_km'].sum()), 1),
    'total_duration_hr': round(float(kpi_source['duration_min'].sum()/60), 1),
    'total_elevation_gain_m': int(kpi_source['elevation_gain_m'].sum()),
    'first_activity': str(kpi_source['date'].min()),
    'last_activity': str(kpi_source['date'].max()),
    'median_distance_km': round(float(kpi_source['distance_km'].median()), 1),
    'median_pace_min_km': round(float(kpi_source['pace_min_km'].median()), 2),
    'average_hr': round(float(kpi_source['avg_hr'].dropna().mean()), 0) if kpi_source['avg_hr'].notna().any() else None,
    'activity_types': ', '.join(sorted(kpi_source['sport'].dropna().unique()))
}
with open(DATA_DIR/'project_kpis.json','w') as f: json.dump(kpis, f, indent=2)

# -------------------------
# AusPlay processing
# -------------------------
BY_SPORT = BASE/'C4S-AusPlay-By-Sport-Data-Tables-30-April-2026.xlsx'
NATIONAL = BASE/'C4S-AusPlay-National-Data-Tables-30-April-2026.xlsx'

wb=load_workbook(BY_SPORT, data_only=True, read_only=True)

# Top adult participation activities from sheet 1 rows 16-157
ws=wb['1']
top=[]
for row in ws.iter_rows(min_row=16, max_row=157, values_only=True):
    activity=row[0]
    rate=row[1]
    if isinstance(activity, str) and isinstance(rate, (int,float)):
        top.append({'activity':activity, 'participation_rate':round(rate*100,1), 'highlight': 'Running/jogging' if activity=='Running/jogging' else 'Other activity'})
top_sorted=sorted(top, key=lambda x: x['participation_rate'], reverse=True)[:12]
# Make sure Running/jogging included if just outside top 12 (it should be in top 12)
if not any(x['activity']=='Running/jogging' for x in top_sorted):
    top_sorted=top_sorted[:-1]+[next(x for x in top if x['activity']=='Running/jogging')]
with open(DATA_DIR/'ausplay_top_activities.json','w') as f: json.dump(top_sorted, f, indent=2)

# Running demographics: rates and counts from sheet 1 and organised sheet 3
labels = {1:'Total adults',3:'Males',4:'Females',6:'15–17',7:'18–24',8:'25–34',9:'35–44',10:'45–54',11:'55–64',12:'65+'}
def running_rows(sheet):
    ws=wb[sheet]
    rows=[]
    for row in ws.iter_rows(values_only=True):
        if row and row[0]=='Running/jogging':
            rows.append(row)
    return rows
r_rate, r_count = running_rows('1')
o_rate, o_count = running_rows('3')
demo=[]
for idx,label in labels.items():
    demo.append({'measure':'Any running/jogging','group':label,'participation_rate':round((r_rate[idx] or 0)*100,1),'estimated_people':round(r_count[idx] or 0)})
    demo.append({'measure':'Organised running/jogging','group':label,'participation_rate':round((o_rate[idx] or 0)*100,1),'estimated_people':round(o_count[idx] or 0)})
with open(DATA_DIR/'ausplay_running_demographics.json','w') as f: json.dump(demo, f, indent=2)

gender=[]
for idx,label in [(3,'Males'),(4,'Females')]:
    gender.append({'measure':'Any running/jogging','gender':label,'participation_rate':round((r_rate[idx] or 0)*100,1),'estimated_people':round(r_count[idx] or 0)})
    gender.append({'measure':'Organised running/jogging','gender':label,'participation_rate':round((o_rate[idx] or 0)*100,1),'estimated_people':round(o_count[idx] or 0)})
with open(DATA_DIR/'ausplay_running_gender.json','w') as f: json.dump(gender, f, indent=2)

# Organising entity by running from sheet 4
ws=wb['4']
entity_header=None
# From manual inspection: columns B-G: total, sports club/association, fitness/leisure centre, other org, school/uni, workplace, other.
entity_names=['Total organised','Sports club/association','Fitness/leisure centre','Other organisation','School/university','Workplace','Other']
rows=[]
for row in ws.iter_rows(values_only=True):
    if row and row[0]=='Running/jogging': rows.append(row)
if rows:
    ent_rate, ent_count = rows[0], rows[1]
    entities=[]
    for col,name in enumerate(entity_names, start=1):
        if col < len(ent_rate) and ent_rate[col] is not None:
            entities.append({'entity':name,'participation_rate':round(ent_rate[col]*100,2),'estimated_people':round(ent_count[col] if col < len(ent_count) and ent_count[col] else 0)})
    with open(DATA_DIR/'ausplay_running_entities.json','w') as f: json.dump(entities, f, indent=2)

# Year comparison from sheets 6 and 7
rows6=running_rows('6'); rows7=running_rows('7')
year_compare=[]
if rows6:
    for year, col in [(2024,1),(2025,3)]:
        year_compare.append({'measure':'Any running/jogging','year':year,'participation_rate':round(rows6[0][col]*100,1),'estimated_people':round(rows6[1][col])})
if rows7:
    for year, col in [(2024,1),(2025,3)]:
        year_compare.append({'measure':'Organised running/jogging (children)','year':year,'participation_rate':round(rows7[0][col]*100,1),'estimated_people':round(rows7[1][col])})
with open(DATA_DIR/'ausplay_running_year_compare.json','w') as f: json.dump(year_compare, f, indent=2)

# State map: National sheet 34, 2025, All, At least three times per week
wb2=load_workbook(NATIONAL, data_only=True, read_only=True)
ws=wb2['34']
state_cols={'ACT':3,'NSW':4,'NT':5,'QLD':6,'SA':7,'TAS':8,'VIC':9,'WA':10,'AUSTRALIA':11}
state_full={'ACT':'Australian Capital Territory','NSW':'New South Wales','NT':'Northern Territory','QLD':'Queensland','SA':'South Australia','TAS':'Tasmania','VIC':'Victoria','WA':'Western Australia','AUSTRALIA':'Australia'}
state_rows=[]
current_group=None; current_year=None
for row in ws.iter_rows(values_only=True):
    if row[0] not in [None, '']:
        current_group=row[0]
    if row[1] not in [None, '']:
        current_year=row[1]
    freq=row[2]
    if current_group=='All' and str(current_year)=='2025' and freq=='At least three times per week':
        for abbr,col in state_cols.items():
            val=row[col]
            state_rows.append({'state_abbr':abbr,'state_name':state_full[abbr], 'frequency':'At least three times per week', 'rate_pct':round(val*100,1)})
with open(DATA_DIR/'ausplay_state_frequency.json','w') as f: json.dump([r for r in state_rows if r['state_abbr']!='AUSTRALIA'], f, indent=2)

# Copy GeoJSON
# Rewind ABS GeoJSON polygons for Vega/D3.
# Without this, Vega-Lite can fill the outside of Australia as a large rectangle.
import json
from shapely.geometry import shape, mapping, MultiPolygon, Polygon
from shapely.geometry.polygon import orient
with open(BASE/'aus_states_territories_2021_simplified.geojson') as f:
    aus_geo = json.load(f)
for feat in aus_geo['features']:
    geom = shape(feat['geometry'])
    if isinstance(geom, Polygon):
        geom = orient(geom, sign=1.0)
    elif isinstance(geom, MultiPolygon):
        geom = MultiPolygon([orient(poly, sign=1.0) for poly in geom.geoms])
    feat['geometry'] = mapping(geom)
with open(DATA_DIR/'aus_states_territories_2021_simplified.geojson', 'w') as f:
    json.dump(aus_geo, f, separators=(',', ':'))

# Sources metadata
sources=[
    {'name':'COROS activity export','file':'exportSportData_466571768635605001_20260518.zip','use':'Personal activity summaries only; raw GPS is not included in the public project data.'},
    {'name':'AusPlay By Sport Data Tables','file':'C4S-AusPlay-By-Sport-Data-Tables-30-April-2026.xlsx','use':'Running/jogging participation rates, rankings and demographics.'},
    {'name':'AusPlay National Data Tables','file':'C4S-AusPlay-National-Data-Tables-30-April-2026.xlsx','use':'State/territory physical activity frequency for the required map.'},
    {'name':'ABS ASGS Edition 3 States and Territories 2021','file':'aus_states_territories_2021_simplified.geojson','use':'Simplified state/territory GeoJSON map boundaries.'}
]
with open(DATA_DIR/'sources.json','w') as f: json.dump(sources, f, indent=2)

# Copy processing script and parser for reproducibility
shutil.copyfile('/mnt/data/fit_minimal.py', SCRIPTS_DIR/'fit_minimal.py')
shutil.copyfile(__file__, SCRIPTS_DIR/'process_data.py')

print('Activities:', len(activities), 'usable summaries')
print('Sports:', adf['sport'].value_counts().to_dict())
print('KPI:', kpis)
print('Data files:', sorted([p.name for p in DATA_DIR.glob('*')]))
